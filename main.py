"""
main.py — India Trade Intelligence (Tradestat Edition)

Put ALL your tradestat Excel files in:  data/raw/tradestat/

Handles all 3 file types automatically:
  tradestat_commodity_export_2023_Jan.xlsx
  tradestat_country_import_2024_Aug.xlsx
  tradestat_cxc_export_2023_Apr_AUSTRALIA.xlsx

Commands:
  python main.py --process     Read all files -> DB -> Excel
  python main.py --status      Show what data exists
  python main.py --excel       Regenerate Excel only
  python main.py --dashboard   Launch Streamlit dashboard
"""

import os, sys, time, logging, argparse
from datetime import datetime
import pandas as pd
import numpy as np

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

BASE         = os.path.dirname(os.path.abspath(__file__))
DATA_DIR     = os.path.join(BASE, "data", "raw", "tradestat")
DB_PATH      = os.path.join(BASE, "data", "trade_tradestat.db")
EXCEL_DIR    = os.path.join(BASE, "data", "excel")
LOG_DIR      = os.path.join(BASE, "data", "logs")

sys.path.insert(0, os.path.join(BASE, "src"))
for d in [DATA_DIR, EXCEL_DIR, LOG_DIR, os.path.join(BASE,"data","processed")]:
    os.makedirs(d, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(os.path.join(LOG_DIR,"trade.log"), mode="a", encoding="utf-8"),
    ]
)
logger = logging.getLogger(__name__)


def cmd_status():
    import glob
    from parser_tradestat import parse_filename
    from database_ts import get_db_stats

    files = glob.glob(os.path.join(DATA_DIR,"*.xlsx")) + glob.glob(os.path.join(DATA_DIR,"*.xls"))
    print(f"\n=== STATUS ===\n")
    print(f"  Data folder : {DATA_DIR}")
    print(f"  Files found : {len(files)}")

    if files:
        types = {"commodity":0,"country":0,"cxc":0}
        years, countries = set(), set()
        for f in files:
            m = parse_filename(f)
            if m["file_type"] in types: types[m["file_type"]] += 1
            if m["year"]: years.add(m["year"])
            if m["country"]: countries.add(m["country"])
        print(f"  Commodity files : {types['commodity']}")
        print(f"  Country files   : {types['country']}")
        print(f"  CXC files       : {types['cxc']} ({len(countries)} countries)")
        print(f"  Years detected  : {sorted(years)}")

    if os.path.exists(DB_PATH):
        s = get_db_stats(DB_PATH)
        print(f"\n  DATABASE:")
        print(f"  CXC rows (HS4 x Country) : {s['cxc_rows']:,}")
        print(f"  Commodity rows           : {s['commodity_rows']:,}")
        print(f"  Country rows             : {s['country_rows']:,}")
        print(f"  HS4 codes                : {s['hs_codes']:,}")
        print(f"  Countries                : {s['countries']:,}")
        print(f"  Date range               : {s['date_range']}")
    else:
        print(f"\n  No database yet. Run: python main.py --process")
    print()


def cmd_process():
    from parser_tradestat import (load_all_files, build_commodity_pivot,
                                   build_country_pivot, build_cxc_pivot, print_summary)
    from database_ts import init_db, upsert_commodity, upsert_country, upsert_cxc, get_db_stats

    start = time.time()
    print(f"\n=== PROCESSING ALL TRADESTAT FILES ===")
    print(f"  Folder: {DATA_DIR}\n")

    data = load_all_files(DATA_DIR)
    print_summary(data)

    if all(df.empty for df in data.values()):
        print("No data found. Check filenames and folder.")
        return

    init_db(DB_PATH)
    totals = {}

    if not data["commodity"].empty:
        w = build_commodity_pivot(data["commodity"])
        n = upsert_commodity(w, DB_PATH)
        totals["commodity"] = n
        print(f"  Stored commodity : {n:,} rows")

    if not data["country"].empty:
        w = build_country_pivot(data["country"])
        n = upsert_country(w, DB_PATH)
        totals["country"] = n
        print(f"  Stored country   : {n:,} rows")

    if not data["cxc"].empty:
        w = build_cxc_pivot(data["cxc"])
        n = upsert_cxc(w, DB_PATH)
        totals["cxc"] = n
        print(f"  Stored CXC       : {n:,} rows  (HS4 x Country data)")

    print(f"\n  Generating Excel ...")
    excel_path = _gen_excel(data)
    if excel_path:
        print(f"  Excel : {excel_path}")

    s = get_db_stats(DB_PATH)
    dur = time.time() - start
    print(f"\n{'='*55}")
    print(f"  DONE in {dur:.0f}s")
    print(f"  CXC rows (HS4 x Country) : {s['cxc_rows']:,}")
    print(f"  HS4 codes                : {s['hs_codes']:,}")
    print(f"  Countries                : {s['countries']:,}")
    print(f"  Date range               : {s['date_range']}")
    print(f"\n  python main.py --dashboard")
    print(f"{'='*55}\n")


def _gen_excel(data: dict) -> str:
    """Generate the 6-sheet Excel workbook."""
    try: import openpyxl
    except: print("  pip install openpyxl"); return None

    from parser_tradestat import build_commodity_pivot, build_country_pivot, build_cxc_pivot

    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(EXCEL_DIR, f"India_Trade_{ts}.xlsx")

    def style(ws, h="#1F4E79"):
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
        hf = PatternFill(start_color=h[1:], end_color=h[1:], fill_type="solid")
        af = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
        for cell in ws[1]:
            cell.fill=hf; cell.font=Font(color="FFFFFF",bold=True,size=9)
            cell.alignment=Alignment(horizontal="center")
        for i,row in enumerate(ws.iter_rows(min_row=2),2):
            for cell in row:
                if i%2==0: cell.fill=af
                cell.font=Font(size=8)
        for ci,col in enumerate(ws.columns,1):
            ml=max((len(str(c.value or "")) for c in col),default=8)
            ws.column_dimensions[get_column_letter(ci)].width=min(ml+2,55)
        ws.freeze_panes="A2"

    with pd.ExcelWriter(path, engine="openpyxl") as writer:

        # Sheet 1: CXC Summary — latest month, all HS4 x Country, sorted by export
        cxc_raw = data.get("cxc", pd.DataFrame())
        if not cxc_raw.empty:
            cxc_w = build_cxc_pivot(cxc_raw)
            latest = cxc_w["date"].max()
            s1 = cxc_w[cxc_w["date"]==latest].copy()
            s1 = s1[s1["export_usd_mn"] > 0].sort_values(
                ["hs_code","export_usd_mn"], ascending=[True,False])
            s1["month_str"] = s1["date"].dt.strftime("%b-%Y")
            out = s1[["month_str","hs_code","hs_description","country",
                       "export_usd_mn","export_share_pct",
                       "import_usd_mn","import_share_pct","trade_balance"]].round(2)
            out.columns = ["Month","HS Code","Description","Country",
                           "Export (USD Mn)","Export % Share",
                           "Import (USD Mn)","Import % Share","Balance (USD Mn)"]
            out.to_excel(writer, sheet_name="HS4 x Country (Latest)", index=False)
            style(writer.sheets["HS4 x Country (Latest)"], "#1A5276")
            logger.info(f"  Sheet 'HS4 x Country': {len(out):,} rows")

        # Sheet 2: CXC All months
        if not cxc_raw.empty:
            cxc_w = build_cxc_pivot(cxc_raw)
            s2 = cxc_w[cxc_w["export_usd_mn"] > 0].copy()
            s2["month_str"] = s2["date"].dt.strftime("%b-%Y")
            out2 = s2[["month_str","year","hs_code","hs_description","country",
                        "export_usd_mn","export_share_pct",
                        "import_usd_mn","import_share_pct"]].round(2)
            out2.columns = ["Month","Year","HS Code","Description","Country",
                            "Export (USD Mn)","Export %","Import (USD Mn)","Import %"]
            out2.to_excel(writer, sheet_name="HS4 x Country (All Months)", index=False)
            style(writer.sheets["HS4 x Country (All Months)"], "#1A5276")

        # Sheet 3: Commodity monthly totals
        comm = data.get("commodity", pd.DataFrame())
        if not comm.empty:
            w = build_commodity_pivot(comm)
            w["month_str"] = w["date"].dt.strftime("%b-%Y")
            out3 = w[["month_str","year","hs_code","hs_description",
                       "export_usd_mn","import_usd_mn","trade_balance"]].round(2)
            out3.columns = ["Month","Year","HS Code","Description",
                            "Export (USD Mn)","Import (USD Mn)","Balance (USD Mn)"]
            out3.to_excel(writer, sheet_name="Commodity Monthly", index=False)
            style(writer.sheets["Commodity Monthly"])

        # Sheet 4: Country monthly totals
        ctry = data.get("country", pd.DataFrame())
        if not ctry.empty:
            w2 = build_country_pivot(ctry)
            w2["month_str"] = w2["date"].dt.strftime("%b-%Y")
            out4 = w2[["month_str","year","country",
                        "export_usd_mn","import_usd_mn","trade_balance"]].round(2)
            out4.columns = ["Month","Year","Country",
                            "Export (USD Mn)","Import (USD Mn)","Balance (USD Mn)"]
            out4.to_excel(writer, sheet_name="Country Monthly", index=False)
            style(writer.sheets["Country Monthly"], "#145A32")

        # Sheet 5: HS4 summary across all months + countries
        if not cxc_raw.empty:
            cxc_w = build_cxc_pivot(cxc_raw)
            s5 = (cxc_w.groupby(["hs_code","hs_description"])
                        .agg(total_export=("export_usd_mn","sum"),
                             total_import=("import_usd_mn","sum"),
                             country_count=("country","nunique"),
                             months=("date","nunique"))
                        .reset_index())
            s5["balance"] = s5["total_export"] - s5["total_import"]
            s5["status"]  = s5["balance"].apply(lambda x: "SURPLUS" if x>=0 else "DEFICIT")
            s5 = s5.sort_values("total_export",ascending=False).round(2)
            s5.columns = ["HS Code","Description","Total Export","Total Import",
                          "# Countries","# Months","Balance","Status"]
            s5.to_excel(writer, sheet_name="HS4 Summary", index=False)
            style(writer.sheets["HS4 Summary"], "#4A235A")

        # Sheet 6: Country summary
        if not cxc_raw.empty:
            cxc_w = build_cxc_pivot(cxc_raw)
            s6 = (cxc_w.groupby("country")
                        .agg(total_export=("export_usd_mn","sum"),
                             total_import=("import_usd_mn","sum"),
                             hs_count=("hs_code","nunique"),
                             months=("date","nunique"))
                        .reset_index())
            s6["balance"]    = s6["total_export"] - s6["total_import"]
            s6["exp_share%"] = (s6["total_export"]/s6["total_export"].sum()*100).round(2)
            s6["imp_share%"] = (s6["total_import"]/s6["total_import"].sum()*100).round(2)
            s6 = s6.sort_values("total_export",ascending=False).round(2)
            s6.columns = ["Country","Total Export","Total Import","# HS Codes",
                          "# Months","Balance","Export Share %","Import Share %"]
            s6.to_excel(writer, sheet_name="Country Summary", index=False)
            style(writer.sheets["Country Summary"], "#145A32")

        # Metadata
        pd.DataFrame([
            ["Generated",datetime.now().strftime("%Y-%m-%d %H:%M")],
            ["Source","Ministry of Commerce — tradestat.commerce.gov.in"],
            ["Values","USD Million (FOB exports, CIF imports)"],
            ["CXC files","One file per country per month — HS4 level breakdown"],
        ],columns=["Field","Value"]).to_excel(writer,sheet_name="Info",index=False)
        style(writer.sheets["Info"],"#2E4053")

    return path


def cmd_excel():
    from parser_tradestat import load_all_files
    data = load_all_files(DATA_DIR)
    path = _gen_excel(data)
    if path: print(f"Excel: {path}")


def cmd_dashboard():
    import subprocess
    app = os.path.join(BASE, "dashboard", "app_ts.py")
    print(f"\nDashboard: http://localhost:8501\n")
    subprocess.run([sys.executable, "-m", "streamlit", "run", app,
                    "--server.port", "8501"])


if __name__ == "__main__":
    p = argparse.ArgumentParser(description="India Trade Intelligence")
    p.add_argument("--process",   action="store_true")
    p.add_argument("--status",    action="store_true")
    p.add_argument("--excel",     action="store_true")
    p.add_argument("--dashboard", action="store_true")
    args = p.parse_args()

    if   args.process:   cmd_process()
    elif args.status:    cmd_status()
    elif args.excel:     cmd_excel()
    elif args.dashboard: cmd_dashboard()
    else:
        print("\nIndia Trade Intelligence — Tradestat Edition")
        print("---------------------------------------------")
        print(f"Data folder: {DATA_DIR}")
        print()
        print("Copy ALL your tradestat Excel files to the data folder above, then:")
        print()
        print("  python main.py --status      Check what files you have")
        print("  python main.py --process     Process everything -> DB + Excel")
        print("  python main.py --dashboard   Launch dashboard")
        print("  python main.py --excel       Regenerate Excel only")
        print()
